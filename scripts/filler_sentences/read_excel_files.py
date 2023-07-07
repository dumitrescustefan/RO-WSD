import os, pathlib, pickle, sys
import openpyxl
import rowordnet as rwn
from datetime import datetime



# Adaugam datele din excel in database.
_INPUT_EXCEL_FILES_ = "scripts/filler_sentences/input/filled_excel/"
_INPUT_DATASET_ = "scripts/filler_sentences/input/dataset/dataset.pickle"
_OUTPUT_DATASET_ = "scripts/filler_sentences/output/dataset/"
_CHECK_ROW_ = 4 # Primele 4 linii din excel contin legenda, divider, nume, divider
_BLOCK_HEIGHT_ = 12 # 1-divider + 1-glosa_literali + 10-propozitii



# Prin read_excel introducem datele din excel intr-o structura de date de forma:
# Dataset.pickle are același format
'''
    {
    "drama" :  [
                    {
                        "user_id" : "nume si prenume" <- din excel
                        "literal" : "dramă"
                        "synsets" : "ENG-... ENG-... ENG-... -1"     <- toate synseturile literalului din wordnet
                        "correct_synset_id" : "ENG-..."      <- synsetul din excel pentru care voluntarii au compus propozitii
                        "text_prefix" : "In romanul sau publicat in 2013, si muntii au ecou, Hosseini povesteste"
                        "text" : "drama"
                        "text_postfix" : "unei fetite pe nume Pari dintr-un sat fictiv."
                        "sentence" : "In romanul sau publicat in 2013, si muntii au ecou, Hosseini povesteste drama unei fetite pe nume Pari dintr-un sat fictiv"
                    },
                    ...
                ]

    "armată : [
                    {
                        ...
                    }
            ]
    ...
}
 '''



# Functia de citire a datelor din excel, si de generare a dictionarului care o sa le contina temporar
def read_excel(save):
    data = {}

    # Load dataset si rwn, folosim rwn pentru a extrage toate synseturile posibile din care face parte un literal
    dataset = {}
    with open(_INPUT_DATASET_, "rb") as f:
        dataset = pickle.load(f)
    wn = rwn.RoWordNet()

    # Luam numele tuturor fisierelor excel si le punem in excel_name_list
    excel_name_list = os.listdir(_INPUT_EXCEL_FILES_)
    for name in excel_name_list:
        extension = pathlib.Path(name).suffix

        # Verificam data fisierul este de tipul xlsx (Excel)
        if extension ==".xlsx":
            temp_name = _INPUT_EXCEL_FILES_ + name
            workbook = openpyxl.load_workbook(filename = temp_name)
            sheet = workbook.active

            #Salvam numele si prenumele
            user_id = sheet.cell(2+1, 2).value
            user_id = user_id.strip()

            # Read data while there exists a literal in (_CHECK_ROW_ + _BLOCK_HEIGHT_*index+1,1)  <--- Verificam spatiul unde ar trebui sa fie literalul,
            # daca este un literal pe acea pozitie, inseamna ca exista intreg "blocul" ce contine, synset_id, sentence_count, propozitiile prescries si propozitiile scrise de voluntari
            index = 0
            while sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index+1, 1).value != "" and sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index+1, 1).value != None and sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index+1, 1).value != "None":

                # Extragem datele din excel in variabile
                literal = sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index+1, 0+1).value
                synsets = " ".join(str(item) for item in wn.synsets(literal=literal))
                synsets += " -1 "
                correct_synset_id = sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index + 2, 1).value
                filled_sentences = sheet.cell(_CHECK_ROW_ + _BLOCK_HEIGHT_*index + 3, 1).value

                # O sa punem propozitiile in liste
                sentences = []
                text_prefixs = []
                text_postfixs = []
                texts = []

                # Despartim fiecare propozitie in parte in text_prefix, text, text_postfix
                # Pentru asta, ne uitam la ultimele filled_sentences din fiecare block
                for i in range(int(_CHECK_ROW_ + _BLOCK_HEIGHT_*(index+1) - filled_sentences), _CHECK_ROW_ + _BLOCK_HEIGHT_*(index+1)):
                    sentence = sheet.cell(i, 2).value

                    # Verificam ca propozitia sa fie scrisa, daca nu e printam un anunt
                    if sentence != "" and sentence !="None" and sentence !=None:
                        sentence = sentence.strip()
                        text_prefixs.append(sentence.split("{")[0])
                        text_postfixs.append(sentence.split("}")[1])
                        texts.append((sentence.split("{")[1]).split("}")[0])
                        sentences.append(text_prefixs[-1] + texts[-1] + text_postfixs[-1])
                    else:
                        print("Utilizatorul: {}, nu a introdus propoziție in {} pe linia {}".format(user_id,name,i))

                # Generam dictionare pentru fiecare propozitie in parte
                # De mentionat ca deoarece am inceput sa citim propozitiile de la finalul blockului la inceput, propozitiile o sa fie scrisa in ordine inversa in dictionar
                for i in range(len(sentences)):
                    if literal not in data:
                        data[literal] = []
                    temp_dict = {}
                    temp_dict["user_id"] = user_id
                    temp_dict["literal"] = literal
                    temp_dict["synsets"] = synsets
                    temp_dict["correct_synset_id"] = correct_synset_id
                    temp_dict["text_prefix"] = text_prefixs[i]
                    temp_dict["text"] = texts[i]
                    temp_dict["text_postfix"] = text_postfixs[i]
                    temp_dict["sentence"] = sentences[i]

                    # Adaugam dictionarul cu datele propozitie in data
                    data[literal].append(temp_dict)
                index += 1

    if save == "dataset":
        save_in_dataset(data, dataset)
    else:
        if save == "separately":
            save_separately(data)




# Functia care combina datasetul existent cu dictionarul generat de read_excel in dataset_H_M_S.pickle
def save_in_dataset(data, dataset):
    for literal in data:
        # De adaugat safety measure sa nu introducem aceasi propozitie de 2 ori daca rulam pentru dataseturi consecutive???
        dataset[literal] += data[literal]
    currentDateAndTime = datetime.now()
    name = "dataset_combined" + currentDateAndTime.strftime("_%H_%M_%S") + ".pickle"
    with open(_OUTPUT_DATASET_ + name, "wb") as f:
        pickle.dump(dataset, f)



# Functia care salveaza dictionarul generat de read_excel in dataset_H_M_S.pickle
def save_separately(data):
    currentDateAndTime = datetime.now()
    name = "dataset_onlyexcel" + currentDateAndTime.strftime("_%H_%M_%S") + ".pickle"
    with open(_OUTPUT_DATASET_ + name, "wb") as f:
        pickle.dump(data, f)



# save = dataset (calls save_in_dataset la final)
# save = separately (calls save_separately la final)
read_excel(save="dataset")




